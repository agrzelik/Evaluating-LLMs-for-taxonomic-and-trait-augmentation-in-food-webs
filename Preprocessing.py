import os
import re
import sys
import subprocess
import argparse
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

#MY FLAGS
llm_optimization_1 = True
aggregation_tests = False

DEFAULT_VERSION = "v1_baseline"

# -----------------------------------------------------------------------------
# Configuration and Path Setup
# -----------------------------------------------------------------------------
_HERE = Path(__file__).resolve().parent
REF_FOLDER = (_HERE / Path("dataset_20260126_ecobase/processed")).resolve()

def get_paths(version: str):
    """Returns input and output paths for a specific version"""
    
    INPUT_ROOTS = {
        "claude": (_HERE / Path(f"LLM features/Original data augmentation/Claude/{version}")).resolve(),
        "gemini": (_HERE / Path(f"LLM features/Original data augmentation/Gemini/{version}")).resolve(),
        "chatgpt": (_HERE / Path(f"LLM features/Original data augmentation/Chatgpt/{version}")).resolve(),
        "qwen": (_HERE / Path(f"LLM features/Original data augmentation/Qwen/{version}")).resolve(),
    }

    OUTPUT_ROOTS = {
        "claude": (_HERE / Path(f"LLM features/Processed/Claude/{version}")).resolve(),
        "gemini": (_HERE / Path(f"LLM features/Processed/Gemini/{version}")).resolve(),
        "chatgpt": (_HERE / Path(f"LLM features/Processed/Chatgpt/{version}")).resolve(),
        "qwen": (_HERE / Path(f"LLM features/Processed/Qwen/{version}")).resolve(),
    }
    
    return INPUT_ROOTS, OUTPUT_ROOTS

MISSING_REPORT_FILENAME = "Missing_Nodes_Report.csv"

# -----------------------------------------------------------------------------
# Column Definitions and Mappings
# -----------------------------------------------------------------------------

if aggregation_tests:
    DESIRED_ORDER_BASE = [ "file_name", "node",  "IsAlive", "included_species_latin", "included_species_english", "data_source_for_included_species", "representative_species_english", "representative_species_latin","data_source_for_representative_species", "latin_name",  "kingdom", 	"phylum",	"class",	"order",	"family", 	"genus", 	"species"]
elif llm_optimization_1:
    DESIRED_ORDER_BASE = [ "file_name", "node",  "IsAlive", "included_species_latin", "included_species_english", "data_source_for_included_species", "representative_species_english", "representative_species_latin", "data_source_for_representative_species", "latin_name", "kingdom",'Kingdom_confidence',	"phylum", 'Phylum_confidence',	"class", 'Class_confidence',"order",'Order_confidence',	"family",'Family_confidence', "genus",'Genus_confidence',"species", 'Species_confidence',]
else:
    DESIRED_ORDER_BASE = [ "file_name", "node", 
        "activity_patterns", "age_at_maturity", "biomass",
        "bioturbation_impact", "body_form", "bycatch_risk", "climate_sensitivity",
        "competition_intensity", "defense_mechanisms", "diet_breadth", "ecosystem_role",
        "export", "fecundity", "feeding_guild", "feeding_strategy",
        "feeding_time_preference", "fishing_pressure", "functional_group",
        "generation_time", "iucn_status", "invasive_potential", "is_alive",
        "larval_stage_type", "life_span", "locomotion_mode", "metabolic_rate",
        "metabolic_strategy", "migration_pattern", "osmotic_regulation",
        "parental_investment", "predator_avoidance_strategy", "preferred_depth_zone",
        "prey_capture_strategy", "reproductive_seasonality", "reproductive_strategy",
        "respiration", "salinity_tolerance", "sensory_adaptations", "sexual_dimorphism",
        "skin_covering", "social_behavior", "spawning_strategy", "substrate_association",
        "symbiotic_associations", "temperature_preference", "trophic_level",
        "typical_length", "water_depth", 
    ]

RAW_VARIANTS = [ "Aggregation Level",	"Included Species",	"Kingdom",	"Phylum",	"Class",	"Order",	"Family",	"Genus",	"Species",
    "Activity Patterns", "Age at Maturity", "Biomass", "Bioturbation Impact",
    "Body Form", "Bycatch Risk", "Climate Sensitivity", "Competition Intensity",
    "Defense Mechanisms", "Diet Breadth", "Ecosystem Role", "Export",
    "Extended Marine-Specific Traits", "Fecundity", "Feeding Guild",
    "Feeding Strategy", "Feeding Time Preference", "Fishing Pressure",
    "Functional Group", "Generation Time", "IUCN Status", "Import",
    "Invasive Potential", "IsAlive", "Larval Stage Type", "Life Span",
    "Locomotion Mode", "Metabolic Rate", "Metabolic Strategy", "Migration Pattern",
    "Osmotic Regulation", "Parental Investment", "Predator Avoidance Strategy",
    "Preferred Depth Zone", "Prey Capture Strategy", "Reproductive Seasonality",
    "Reproductive Strategy", "Respiration", "Salinity Tolerance",
    "Sensory Adaptations", "Sexual Dimorphism", "Skin Covering", "Social Behavior",
    "Spawning Strategy", "Substrate Association", "Symbiotic Associations",
    "Temperature Preference", "TrophicLevel", "TrophyLevel", "Typical Length",
    "Water Depth", "file_name", "node",
]

# -----------------------------------------------------------------------------
# Unit Conversion Constants
# -----------------------------------------------------------------------------
TIME_FACTORS_TO_YEARS = {
    "year": 1.0, "years": 1.0, "y": 1.0, "yr": 1.0, "yrs": 1.0,
    "month": 1.0 / 12.0, "months": 1.0 / 12.0, "mo": 1.0 / 12.0,
    "week": 1.0 / 52.1429, "day": 1.0 / 365.2422,
    "hour": 1.0 / 8760.0, "decade": 10.0,
}

LENGTH_FACTORS_TO_CM = {
    "cm": 1.0, "mm": 0.1, "m": 100.0,
    "um": 0.0001, "inch": 2.54, "ft": 30.48,
}

TIME_COLS = ["life_span", "age_at_maturity", "generation_time"]
LENGTH_COLS = ["typical_length"]
METABOLIC_COL = "metabolic_rate"

# -----------------------------------------------------------------------------
# Mapping Logic
# -----------------------------------------------------------------------------
def build_column_mapping() -> Dict[str, str]:
    target_lookup = {re.sub(r"[^a-z0-9]", "", c.lower()): c for c in DESIRED_ORDER_BASE}

    overrides = {
        "trophylevel": "trophic_level",
        "trophiclevel": "trophic_level",
        "isalive": "is_alive",
        "import": "Import",
        "extendedmarinespecifictraits": "extended_marine_specific_traits",
        "osmbioticregulation": "osmotic_regulation",
        "substratesociation": "substrate_association",
        "lifespan": "life_span",
    }

    synonym_groups = {
        ""
        "aggregation_level":['Aggregation Level','Aggregation_Level','aggregation_level','AggregationLevel', 'Aggregation Type', 'AggregationType', 'Aggregation_Explanation', 'Aggregation_Type',],
        'Aggregation Level_confidence':['Aggregation Level_confidence', 'Aggregation_Level_confidence'],
        "included_species_latin":['Included Species Latin','Included Species','Included Species (Latin)','Included_Species_Latin','Included_Species','IncludedSpecies_Latin','included_species_latin'],
        "included_species_english":['Included Species English','Included Species','Included Species (English)',"Data source for 'Included species' column", 'Included_Species_English','Included_Species','IncludedSpecies_English','included_species_english'],
        "data_source_for_included_species":['Included Species Data Source',"Data source for 'Included species'",'Included_Species_Source','Included_Species_Links','Included_Species_Sources','included_species_source','Included_Species_source',
                                            'Included_Species_DataSource','IncludedSpecies_Source','included_species_source',"Data source for 'Included species' column","2b. Data source for 'Included species' column in a form of a link","2b. Data source for 'Included species'","Data source for 'Included species' column"],
        "representative_species_latin": ["Representative Species Latin","Representative Species","Representative Species (Latin)","Representative_Species_Latin","Representative_Species","RepresentativeSpecies_Latin","representative_species_latin"],
        "representative_species_english": ["Representative Species English","Representative Species","Representative Species (English)","Representative_Species_English","Representative_Species","RepresentativeSpecies_English","representative_species_english"],
        "data_source_for_representative_species":['Representative species Data Source',"Data source for 'Representative species'",'Representative_Species_Source','Representative Species Data Source','Representative_Species_Link',
                                                    'representative_species_source','Representative_Species_source','Representative_Species_DataSource','RepresentativeSpecies_Source',"3b. Data source for 'Representative species' column in a form of a link","3b. Data source for 'Representative species'",],
        "latin_name":['Latin Name', 'Latin name of the species', 'LatinName', 'Latin_Name', 'Latin_name',],
        "kingdom":["kingdom","Kingdom", 'tax_kingdom', 'taxon_kingdom'],
        "phylum":["phylum","Phylum", 'tax_phylum', 'taxon_phylum'],
        "class":['Class','class','tax_class', 'taxon_class'],
        "order":["order","Order", 'tax_order', 'taxon_order'],
        "family":["family","Family", 'tax_family', 'taxon_family'],
        "genus":["genus","Genus", 'tax_genus', 'taxon_genus'],
        "species":["species","Species", 'taxon_species', 'tax_species'],
        "activity_patterns": ["Activity Patterns", "ActivityPattern", "ActivityPatterns", "Activity_Patterns", "activityPatterns", "activity_patterns", "activity_pattern"],
        "age_at_maturity": ["Age at Maturity", "AgeAtMaturity", "Age_at_Maturity", "ageAtMaturity", "ageAtMatury", "age_at_maturity", "age_at_maturity_years"],
        "biomass": ["Biomass"],
        "bioturbation_impact": ["Bioturbation", "Bioturbation Impact", "BioturbationImpact", "Bioturbation_Impact", "bioturbationImpact", "bioturbation_impact"],
        "body_form": ["Body Form", "BodyForm", "Body_Form", "bodyForm", "body_form"],
        "bycatch_risk": ["Bycatch Risk", "BycatchRisk", "Bycatch_Risk", "bycatchRisk", "bycatch_risk"],
        "climate_sensitivity": ["Climate Sensitivity", "ClimateSensitivity", "Climate_Sensitivity", "climateSensitivity", "climate_sensitivity"],
        "competition_intensity": ["Competition Intensity", "CompetitionIntensity", "Competition_Intensity", "competitionIntensity", "competition_intensity"],
        "defense_mechanisms": ["Defense Mechanisms", "DefenseMechanisms", "Defense_Mechanisms", "defenseMechanisms", "defense_mechanisms"],
        "diet_breadth": ["Diet Breadth", "DietBreadth", "Diet_Breadth", "dietBreadth", "diet_breadth"],
        "ecosystem_role": ["Ecosystem Role", "EcosystemRole", "Ecosystem_Role", "ecosystemRole", "ecosystem_role"],
        "export": ["Export"],
        "fecundity": ["Fecundity", "fecundity"],
        "feeding_guild": ["Feeding Guild", "Feeding_Guild", "FeedingGuild", "feedingGuild", "feeding_guild"],
        "feeding_strategy": ["Feeding Strategy", "Feeding_Strategy", "FeedingStrategy", "feedingStrategy", "feeding_strategy"],
        "feeding_time_preference": ["Feeding Time Preference", "Feeding_Time_Preference", "FeedingTimePreference", "feedingTimePreference", "feeding_time_preference"],
        "fishing_pressure": ["Fishing Pressure", "Fishing_Pressure", "FishingPressure", "fishingPressure", "fishing_pressure"],
        "functional_group": ["Functional Group", "Functional_Group", "FunctionalGroup", "functionalGroup", "functional_group"],
        "generation_time": ["Generation Time", "Generation_Time", "GenerationTime", "generationTime", "generation_time", "generation_time_years"],
        "iucn_status": ["IUCN Status", "IUCN_Status", "IUCNStatus", "IUCN_status", "iucnStatus", "iucn_status"],
        "invasive_potential": ["Invasive Potential", "Invasive_Potential", "InvasivePotential", "invasivePotential", "invasive_potential"],
        "is_alive": ["IsAlive"],
        "larval_stage_type": ["Larval Stage Type", "Larval_Stage_Type", "LarvalStageType", "larvalStageType", "larval_stage_type"],
        "life_span": ["Life Span", "Life_Span", "LifeSpan", "lifeSpan", "life_span", "lifespan", "life_span_years"],
        "locomotion_mode": ["Locomotion Mode", "Locomotion_Mode", "LocomotionMode", "locomotionMode", "locomotion_mode"],
        "metabolic_rate": ["Metabolic Rate", "Metabolic_Rate", "MetabolicRate", "metabolicRate", "metabolic_rate", "metabolic_rate_mgO2_g_h", "metabolic_rate_mg_O2_g_h"],
        "metabolic_strategy": ["Metabolic Strategy", "Metabolic_Strategy", "MetabolicStrategy", "metabolicStrategy", "metabolic_strategy"],
        "migration_pattern": ["Migration Pattern", "Migration_Pattern", "MigrationPattern", "migrationPattern", "migration_pattern"],
        "osmotic_regulation": ["Osmotic Regulation", "Osmotic_Regulation", "OsmoticRegulation", "osmoticRegulation", "osmotic_regulation", "Osmbiotic_Regulation"],
        "parental_investment": ["Parental Investment", "Parental_Investment", "ParentalInvestment", "parentalInvestment", "parental_investment"],
        "predator_avoidance_strategy": ["Predator Avoidance Strategy", "Predator_Avoidance_Strategy", "PredatorAvoidanceStrategy", "predatorAvoidanceStrategy", "predator_avoidance_strategy"],
        "preferred_depth_zone": ["Preferred Depth Zone", "Preferred_Depth_Zone", "PreferredDepthZone", "preferredDepthZone", "preferred_depth_zone"],
        "prey_capture_strategy": ["Prey Capture Strategy", "Prey_Capture_Strategy", "PreyCaptureStrategy", "preyCaptureStrategy", "prey_capture_strategy"],
        "reproductive_seasonality": ["Reproductive Seasonality", "Reproductive_Seasonality", "ReproductiveSeasonality", "reproductiveSeasonality", "reproductive_seasonality"],
        "reproductive_strategy": ["Reproductive Strategy", "Reproductive_Strategy", "ReproductiveStrategy", "reproductiveStrategy", "reproductive_strategy"],
        "respiration": ["Respiration"],
        "salinity_tolerance": ["Salinity Tolerance", "Salinity_Tolerance", "SalinityTolerance", "salinityTolerance", "salinity_tolerance"],
        "sensory_adaptations": ["Sensory Adaptations", "Sensory_Adaptations", "SensoryAdaptations", "sensoryAdaptations", "sensory_adaptations"],
        "sexual_dimorphism": ["Sexual Dimorphism", "Sexual_Dimorphism", "SexualDimorphism", "sexualDimorphism", "sexual_dimorphism"],
        "skin_covering": ["Skin Covering", "Skin_Covering", "SkinCovering", "skinCovering", "skin_covering"],
        "social_behavior": ["Social Behavior", "Social_Behavior", "SocialBehavior", "socialBehavior", "social_behavior"],
        "spawning_strategy": ["Spawning Strategy", "Spawning_Strategy", "SpawningStrategy", "spawningStrategy", "spawning_strategy"],
        "substrate_association": ["Substrate Association", "Substrate_Association", "SubstrateAssociation", "substrateAssociation", "substrateSociation", "substrate_association"],
        "symbiotic_associations": ["Symbiotic Associations", "Symbiotic_Associations", "SymbioticAssociations", "symbioticAssociations", "symbiotic_associations"],
        "temperature_preference": ["Temperature Preference", "Temperature_Preference", "TemperaturePreference", "temperaturePreference", "temperature_preference"],
        "trophic_level": ["TrophicLevel", "TrophyLevel"],
        "typical_length": ["Typical Length", "Typical_Length", "TypicalLength", "typicalLength", "typical_length", "typical_length_cm"],
        "water_depth": ["Water Depth", "Water_Depth", "WaterDepth", "waterDepth", "water_depth"],
        "file_name": ["file_name"],
        "node": ["node"],
        "Import": ["Import"],
    }

    mapping = {}

    def norm(s: str) -> str:
        return re.sub(r"[^a-z0-9]", "", s.lower())

    for canonical, variants in synonym_groups.items():
        canonical_target = (
            canonical if canonical in DESIRED_ORDER_BASE else overrides.get(norm(canonical), canonical)
        )
        resolved = target_lookup.get(
            norm(canonical_target), overrides.get(norm(canonical_target), canonical_target)
        )
        for v in variants:
            mapping[v] = resolved

    for k, v in overrides.items():
        mapping[k] = v

    for raw in RAW_VARIANTS:
        n = norm(str(raw))
        if n in target_lookup:
            mapping[raw] = target_lookup[n]
        else:
            cleaned = n.replace("years", "").replace("cm", "")
            if cleaned in target_lookup:
                mapping[raw] = target_lookup[cleaned]

    return mapping


COLUMN_MAP = build_column_mapping()

# -----------------------------------------------------------------------------
# Data Parsing and Cleaning Functions
# -----------------------------------------------------------------------------
def normalize_range(value) -> Optional[str]:
    if pd.isna(value) or str(value).strip().lower() in ["", "nan", "none", "not available"]:
        return None
    s = str(value).strip()
    s = s.replace("â€“", "-").replace("–", "-").replace("—", "-").replace("−", "-")
    s = re.sub(r"(?<=\d)\s*(?:to)\s*(?=\d)", "-", s, flags=re.IGNORECASE)
    return re.sub(r"\s*-\s*", "-", s)


def clean_metabolic_rate_string(text) -> Optional[str]:
    if pd.isna(text):
        return None
    s = normalize_range(text)
    if not s:
        return None
    match = re.match(r"^[\d\.]+(?:\s*-\s*[\d\.]+)?", s)
    return match.group(0) if match else None


def _parse_numbers(text: str) -> List[float]:
    nums = re.findall(r"\d+(?:\.\d+)?", text)
    return [float(n) for n in nums]


def parse_time_to_years(text) -> Optional[float]:
    if pd.isna(text):
        return None
    t = normalize_range(text)
    if not t:
        return None
    tl = t.lower()
    units_found = [u for u in TIME_FACTORS_TO_YEARS if re.search(rf"\b{re.escape(u)}\b", tl)]
    dominant = max(units_found, key=lambda u: TIME_FACTORS_TO_YEARS[u]) if units_found else "year"
    nums = _parse_numbers(tl)
    if not nums:
        return None
    conv = [n * TIME_FACTORS_TO_YEARS.get(dominant, 1.0) for n in nums]
    return sum(conv) / len(conv)


def parse_length_to_cm(text) -> Optional[float]:
    if pd.isna(text):
        return None
    t = normalize_range(text)
    if not t:
        return None
    units_found = [u for u in LENGTH_FACTORS_TO_CM if re.search(rf"\b{re.escape(u)}\b", t.lower())]
    dominant = max(units_found, key=lambda u: LENGTH_FACTORS_TO_CM[u]) if units_found else "cm"
    nums = _parse_numbers(t)
    if not nums:
        return None
    conv = [n * LENGTH_FACTORS_TO_CM.get(dominant, 1.0) for n in nums]
    return sum(conv) / len(conv)


def fmt_num(x: Optional[float]) -> Optional[str]:
    if x is None:
        return None
    return re.sub(r"\.0+$", "", f"{x:.6f}")


def parse_metabolic_rate_to_mean(text) -> Optional[float]:
    if pd.isna(text):
        return None
    s = clean_metabolic_rate_string(text)
    if not s:
        return None
    nums = _parse_numbers(s)
    if not nums:
        return None
    return sum(nums) / len(nums)


def _normalize_key(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())

# -----------------------------------------------------------------------------
def _map_record_columns(rec: Dict[str, object], filename: str) -> Dict[str, object]:
    """
    Maps the raw columns of a single record (row) to the canonical schema
    using the global COLUMN_MAP.
    """
    new_rec: Dict[str, object] = {"file_name": filename}

    for orig_col, val in rec.items():
        if orig_col in COLUMN_MAP:
            target_col = COLUMN_MAP[orig_col]
            if target_col not in new_rec or pd.isna(new_rec.get(target_col)):
                if not pd.isna(val):
                    new_rec[target_col] = val
        else:
            # Fallback normalization check
            norm_orig = _normalize_key(str(orig_col))
            for t in DESIRED_ORDER_BASE:
                if _normalize_key(t) == norm_orig:
                    if not pd.isna(val):
                        new_rec[t] = val
                    break

    # Fallback to finding the 'node' if missing
    if "node" not in new_rec or pd.isna(new_rec.get("node")):
        try:
            first = list(rec.values())[0]
            if isinstance(first, str) and len(first) > 1:
                new_rec["node"] = first
        except Exception:
            pass

    # Normalize specific data fields
    for col in TIME_COLS + LENGTH_COLS + [METABOLIC_COL]:
        if col in new_rec:
            new_rec[col] = normalize_range(new_rec[col])

    return new_rec


def _aggregate_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Performs unit conversions and aggregations for time, length, 
    and metabolic rate columns.
    """
    out = df.copy()

    for col in TIME_COLS:
        if col in out.columns:
            out[f"{col}_avg_years"] = out[col].apply(parse_time_to_years).apply(fmt_num)
            out.drop(columns=[col], inplace=True)

    for col in LENGTH_COLS:
        if col in out.columns:
            out[f"{col}_avg_cm"] = out[col].apply(parse_length_to_cm).apply(fmt_num)
            out.drop(columns=[col], inplace=True)

    if METABOLIC_COL in out.columns:
        out[f"{METABOLIC_COL}_avg"] = out[METABOLIC_COL].apply(parse_metabolic_rate_to_mean).apply(fmt_num)
        out.drop(columns=[METABOLIC_COL], inplace=True)

    return out


def _dedupe_and_fill(df: pd.DataFrame) -> pd.DataFrame:
    """
    Removes duplicate records based on file_name and node. 
    Prioritizes records with higher completeness before filling missing values.
    """
    out = df.copy()
    out["completeness"] = out.notna().sum(axis=1)
    out.sort_values(["file_name", "node", "completeness"], ascending=[True, True, False], inplace=True)
    out.drop_duplicates(subset=["file_name", "node"], keep="first", inplace=True)
    out.drop(columns=["completeness"], inplace=True)
    return out.fillna("NA")


def _final_order(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
    """
    Reorders DataFrame columns to match the desired output structure,
    inserting missing columns with default values.
    """
    final_ordered_cols: List[str] = []
    for c in DESIRED_ORDER_BASE:
        if c in TIME_COLS:
            final_ordered_cols.append(f"{c}_avg_years")
        elif c in LENGTH_COLS:
            final_ordered_cols.append(f"{c}_avg_cm")
        elif c == METABOLIC_COL:
            final_ordered_cols.append(f"{c}_avg")
        else:
            final_ordered_cols.append(c)

    out = df.copy()
    for c in final_ordered_cols:
        if c not in out.columns:
            out[c] = "NA"

    return out[final_ordered_cols], final_ordered_cols


def _save_output(df: pd.DataFrame, output_folder: Path, output_filename: str) -> Path:
    """Saves the processed DataFrame to an Excel file."""
    output_folder.mkdir(parents=True, exist_ok=True)
    out_path = output_folder / output_filename
    df.to_excel(out_path, index=False)
    print(f"Saved processed file to: {out_path}")
    return out_path


def verify_nodes_against_flows(gemini_df: pd.DataFrame, ref_folder: Path, output_folder: Path) -> None:
    """
    Verifies if the nodes in the processed DataFrame exist in the reference 
    'Flows' sheets. Generates a report if nodes are missing.
    """
    print(f"\n--- Verifying against 'Flows' in {ref_folder} ---")
    if not ref_folder.exists():
        return

    flows_nodes = set()
    for f in os.listdir(ref_folder):
        if f.endswith(".xlsx"):
            try:
                wb = openpyxl.load_workbook(ref_folder / f, read_only=True, data_only=True)
                if "Flows" in wb.sheetnames:
                    rows = wb["Flows"].iter_rows(values_only=True)
                    headers = next(rows, None)
                    if headers:
                        idxs = [
                            i for i, h in enumerate(headers)
                            if h and re.search(r"(node|source|target|consumer|resource|predator|prey)", str(h).lower())
                        ]
                        for r in rows:
                            for i in idxs:
                                if i < len(r) and r[i]:
                                    flows_nodes.add(str(r[i]).strip())
                wb.close()
            except Exception:
                pass

    gemini_nodes = set(gemini_df["node"].dropna().astype(str).str.strip())
    missing = flows_nodes - gemini_nodes
    
    print(f"Missing Nodes: {len(missing)}")
    if missing:
        pd.DataFrame({"Missing_Nodes": sorted(list(missing))}).to_csv(
            output_folder / MISSING_REPORT_FILENAME, index=False
        )
    else:
        print("All reference nodes found.")


# -----------------------------------------------------------------------------
# Standardize Filenames
# -----------------------------------------------------------------------------
def _standardize_filename(filename: str) -> str:
    """
    Cleans the filename by removing extensions, version numbers (v5, v1, etc.),
    and model names (openai, gemini, etc.) to ensure consistency across models.
    Example: 'Alaska_Prince_William_Sound_2 v5 openai.xlsx' -> 'Alaska_Prince_William_Sound_2'
    """
    # 1. Remove file extension
    name = Path(filename).stem
    
    # 2. Define Regex patterns to remove. 
    patterns_to_remove = [
        r"(?i)[_\s]*v\s*\d+",       
        r"(?i)[_\s]*v\d+_\w+",       
        r"(?i)[_\s]*openai",         
        r"(?i)[_\s]*chatgpt", 
        r"(?i)[_\s]*gemini",         
        r"(?i)[_\s]*anthropic",     
        r"(?i)[_\s]*claude",     
        r"(?i)[_\s]*qwen",           
        r"(?i)[_\s]*deepseek"       
    ]
    
    for pattern in patterns_to_remove:
        name = re.sub(pattern, "", name)
        
    return name.strip()


def process_model_folder(model_name: str, input_folder: Path, output_folder: Path, output_filename: str) -> None:
    """
    Orchestrates the processing for a specific model folder: 
    read -> clean_filename -> map -> aggregate -> dedupe -> save -> verify.
    """
    print(f"--- Processing {model_name} Files ---")
    if not input_folder.exists():
        print(f"Input folder not found: {input_folder}")
        return

    files = [f for f in os.listdir(input_folder) if f.lower().endswith((".xlsx", ".xls"))]
    print(f"Found {len(files)} files.")

    all_records: List[Dict[str, object]] = []
    for filename in files:
        file_path = input_folder / filename
        try:
            # --- Standardize the filename before mapping ---
            clean_name = _standardize_filename(filename)
            
            df = pd.read_excel(file_path)
            for rec in df.to_dict("records"):
                all_records.append(_map_record_columns(rec, clean_name))
        except Exception as e:
            print(f"Error processing {filename}: {e}")

    if not all_records:
        print("No records extracted.")
        return

    final_df = pd.DataFrame(all_records)
    print(f"Total records extracted: {len(final_df)}")

    final_df = _aggregate_columns(final_df)
    final_df = _dedupe_and_fill(final_df)
    final_df, _ = _final_order(final_df)

    _save_output(final_df, output_folder, output_filename)
    verify_nodes_against_flows(final_df, REF_FOLDER, output_folder)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Preprocess LLM augmented data")
    parser.add_argument(
        "--version",
        default=DEFAULT_VERSION,
        help=f"Version to process (default: {DEFAULT_VERSION}). Example: v2_temperature_0.5"
    )
    parser.add_argument(
        "--provider",
        choices=["claude", "gemini", "chatgpt", "qwen", "all"],
        default="all",
        help="Which provider to process (default: all)"
    )
    parser.add_argument(
        "--skip-json-parser",
        action="store_true",
        help="Skip running Jsonparserforllm.py (assume inputs already exist)"
    )
    
    args = parser.parse_args()
    
    print("\n" + "=" * 70)
    print("PREPROCESSING PIPELINE")
    print("=" * 70)
    print(f"Version: {args.version}")
    print(f"Provider: {args.provider}")
    print("=" * 70 + "\n")
    
    # -------------------------------------------------------------------------
    # Step 0: Run Jsonparserforllm.py to ensure input files exist
    # -------------------------------------------------------------------------
    if not args.skip_json_parser:
        print("Running Jsonparserforllm.py to prepare input files...")
        print("=" * 60)
        
        json_parser_script = _HERE / "Jsonparserforllm.py"
        if json_parser_script.exists():
            try:
                # Pass --version to Jsonparser
                subprocess.run(
                    [sys.executable, str(json_parser_script), "--version", args.version, "--provider", args.provider], 
                    check=True
                )
                print("\nJsonparserforllm.py completed successfully.\n")
            except subprocess.CalledProcessError as e:
                print(f"\nError running Jsonparserforllm.py: {e}")
                print("Stopping pipeline.")
                sys.exit(1)
        else:
            print(f"Warning: {json_parser_script} not found. Assuming inputs exist.")
    else:
        print("Skipping Jsonparserforllm.py (--skip-json-parser flag set)\n")

    # -------------------------------------------------------------------------
    # Step 1: Get paths for this version
    # -------------------------------------------------------------------------
    INPUT_ROOTS, OUTPUT_ROOTS = get_paths(args.version)
    
    # -------------------------------------------------------------------------
    # Step 2: Run Preprocessing Logic
    # -------------------------------------------------------------------------
    models_to_process = ["claude", "gemini", "chatgpt", "qwen"] if args.provider == "all" else [args.provider]
    
    for model in models_to_process:
        in_folder = INPUT_ROOTS[model]
        out_folder = OUTPUT_ROOTS[model]
        out_folder.mkdir(parents=True, exist_ok=True)
        out_filename = f"{model}_Processed_Final.xlsx"
        process_model_folder(model, in_folder, out_folder, out_filename)
    
    print("\n" + "=" * 70)
    print("PREPROCESSING COMPLETE!")
    print(f"Outputs saved to: LLM features/Processed/{model}/{args.version}/")
    print("=" * 70)